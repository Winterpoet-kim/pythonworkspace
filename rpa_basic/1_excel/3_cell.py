from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "DalbitSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # A1 셀의 객체 정보를 출력
print(ws["A1"].value)  # A1 셀의 정보를 출력
print(ws["A10"].value)  # 값이 없을땐 None 을 출력

print(ws.cell(row=1, column=1).value)  # == print(ws["A1"].value)
print(ws.cell(row=1, column=2).value)  # == print(ws["B1"].value)

c1 = ws.cell(column=3, row=1, value=10)  # ws["C1"].value = 10
print(c1.value)

from random import *

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))  # 0~100 사이의 숫자

wb.save("sample.xlsx")