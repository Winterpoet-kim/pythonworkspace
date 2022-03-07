from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])
for i in range(1, 11):  # 10개의 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]  # 영어 column만 가지고 오기
# print(col_B)
# for cell in col_B :
#     print(cell.value)


col_range = ws["B:C"]  # B 부터 C 까지
# print(col_range)
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1]  # 첫번째 row 만 가져오기
# for cell in row_title:
#     print(cell.value, end=" ")
    
print()

# row_range = ws[2:6]  # 1번째줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row]  # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         #print(cell.value, end=" ")
#         #print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         #print(xy, end=" ")
#         print(xy[0], end="")  # A, ...
#         print(xy[1], end=" ")  # 1, ...
#     print()


# 전제 rows
#print(tuple(ws.rows))  # 한 줄씩 가져와서 tuple로 만듬
# for row in tuple(ws.rows):
#     print(row[1].value)  # 튜플의 2번째 index의 value 가져오기

# 전제 columns
#print(tuple(ws.columns))  # 한 열씩 가져와서 tuple로 만듬
# for column in tuple(ws.columns):
#     print(column[0].value)  # 튜플의 첫번째 index의 value 가져오기


# 2~11번째줄, 2~3번째열까지
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):  # 범위 지정이 가능함.
#     #print(row[0].value, row[1].value)
#     print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):  # 범위 지정이 가능함.
    #print(row[0].value, row[1].value)
    print(col)




wb.save("sample.xlsx")
