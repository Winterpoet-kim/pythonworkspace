from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet()  # 새로운 sheet 를 기본이름으로 생성
ws.title = "Mysheet"  # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
# Sheet, MySheet, NewSheet, YourSheet
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"]  # Dictionary 형태로 Sheet에 접근 가능

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames)  # 모든 Sheet 이름 확인

wb.save("sample.xlsx")

 