from openpyxl import load_workbook
wb = load_workbook("scores.xlsx")
ws = wb.active

for i in range(2, 12):
    # quiz2 점수 수정
    ws.cell(column=4, row=i, value=10)

    # 총점 수식 넣기
    ws["H1"] = "총점"
    ws[f"H{i}"] = f"=SUM(B{i}:G{i})"



# 성적 부여
ws["I1"] = "성적"
for x in range(2, 12):
    total = 0
    for y in range(2, 8):
        score = ws.cell(row=x, column=y).value
        total += int(score)
        
    if total < 70:
        ws[f"I{x}"] = "D"
    elif 70 <= total < 80:
        ws[f"I{x}"] = "C"
    elif 80 <= total < 90:
        ws[f"I{x}"] = "B"
    else:
        ws[f"I{x}"] = "A"

    # 출석 점수 5 미만 = F
    if int(ws[f"B{x}"].value) < 5:
        ws[f"I{x}"] = "F"    

wb.save("scores_result.xlsx")


